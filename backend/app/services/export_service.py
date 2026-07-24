import os
import json
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.core.database import get_db_connection

def apply_header_formatting(ws, headers):
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    ws.freeze_panes = "A2"
    
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

def auto_size_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        adjusted_width = min(adjusted_width, 60)
        col_letter = column[0].column_letter
        ws.column_dimensions[col_letter].width = adjusted_width

def create_empty_state_message(ws, headers, message="No data available for the selected filters."):
    if len(headers) > 0:
        apply_header_formatting(ws, headers)
    ws.append([message])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(headers)))
    cell = ws.cell(row=2, column=1)
    cell.font = Font(italic=True, color="666666")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    auto_size_columns(ws)

def generate_export_workbook(account_id: str, service_name: str, region_filter: str = None, rec_filter: str = None) -> io.BytesIO:
    conn = get_db_connection()
    wb = Workbook()
    
    # --- Sheet 1: Executive Summary ---
    ws1 = wb.active
    ws1.title = "Executive Summary"
    
    account_name = "Unknown"
    curr = conn.cursor()
    curr.execute("SELECT account_name FROM cloud_configs WHERE account_id = ?", (account_id,))
    row = curr.fetchone()
    if row:
        account_name = row['account_name']
        
    curr.execute("""
        SELECT start_time, end_time, duration_seconds, status 
        FROM pipeline_executions 
        WHERE account_id = ? AND service_name = ? 
        ORDER BY start_time DESC LIMIT 1
    """, (account_id, service_name))
    pipeline = curr.fetchone()
    
    pl_start = pipeline['start_time'] if pipeline else "N/A"
    pl_end = pipeline['end_time'] if pipeline else "N/A"
    pl_duration = f"{pipeline['duration_seconds']}s" if pipeline and pipeline['duration_seconds'] else "N/A"
    pl_status = pipeline['status'] if pipeline else "N/A"
    
    region_clause = ""
    params = [account_id, service_name]
    if region_filter and region_filter != "All Regions":
        region_clause = " AND region = ?"
        params.append(region_filter)
        
    curr.execute(f"""
        SELECT COUNT(*) FROM (
            SELECT DISTINCT region FROM discovered_resources WHERE account_id = ? AND service_type = ? {region_clause}
            UNION
            SELECT DISTINCT region FROM billing_service_cache WHERE account_id = ? AND service_name = ? {region_clause}
        )
    """, params + params)
    res_reg_count = curr.fetchone()[0]
    
    curr.execute(f"SELECT COUNT(resource_id) FROM discovered_resources WHERE account_id = ? AND service_type = ? {region_clause}", params)
    total_discovered = curr.fetchone()[0]
    
    curr.execute(f"SELECT COUNT(resource_id) FROM resource_summaries WHERE account_id = ? AND service_type = ? {region_clause}", params)
    total_analyzed = curr.fetchone()[0]
    
    curr.execute(f"SELECT recommendation FROM resource_summaries WHERE account_id = ? AND service_type = ? {region_clause}", params)
    all_recs = curr.fetchall()
    
    total_recs_generated = 0
    total_upsize = 0
    total_downsize = 0
    total_keep = 0
    total_no_metrics = 0
    total_specific_instance = 0
    
    for r in all_recs:
        rec_text = r['recommendation']
        if rec_text == "Unknown" or rec_text == "Analysis Failed":
            total_no_metrics += 1
        else:
            total_recs_generated += 1
            if "Upsize" in rec_text:
                total_upsize += 1
            elif "Downsize" in rec_text:
                total_downsize += 1
            elif "Recommend Specific Instance" in rec_text or "specific instance" in rec_text.lower():
                total_specific_instance += 1
            else:
                total_keep += 1
                
    now_str = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    
    ws1.append(["Right-Sizing Executive Summary"])
    ws1.append([])
    ws1.append(["AWS Account ID", account_id])
    ws1.append(["Account Name", account_name])
    ws1.append(["Service Name", service_name])
    ws1.append(["Analysis Date & Time", now_str])
    ws1.append(["Lookback Window", "30 Days"]) 
    ws1.append(["Pipeline Start Time", pl_start])
    ws1.append(["Pipeline End Time", pl_end])
    ws1.append(["Total Pipeline Duration", pl_duration])
    ws1.append(["Pipeline Status", pl_status])
    ws1.append([])
    ws1.append(["Analysis Summary"])
    ws1.append(["Total Regions", res_reg_count or 0])
    ws1.append(["Total Resources Discovered", total_discovered or 0])
    ws1.append(["Total Resources Analysed", total_analyzed or 0])
    ws1.append(["Total Recommendations Generated", total_recs_generated])
    ws1.append(["Total Upsize Recommendations", total_upsize])
    ws1.append(["Total Downsize Recommendations", total_downsize])
    ws1.append(["Total Keep Current Recommendations", total_keep])
    ws1.append(["Total Recommend Specific Instance", total_specific_instance])
    ws1.append(["Total Resources Without Metrics", total_no_metrics])
    
    for row_idx in [1, 12]:
        cell = ws1.cell(row=row_idx, column=1)
        cell.font = Font(bold=True, size=14)
    for row in ws1.iter_rows(min_row=3, max_row=10):
        row[0].font = Font(bold=True)
    for row in ws1.iter_rows(min_row=13, max_row=20):
        row[0].font = Font(bold=True)
    auto_size_columns(ws1)


    # --- Sheet 2: Regions Summary ---
    ws2 = wb.create_sheet(title="Regions Summary")
    headers2 = [
        "Region", "Region Processing Status", "Pipeline Execution Status", "Resources Discovered", "Resources Analysed", 
        "Recommendations Generated", "Resources Without Metrics", "Discovery Duration", 
        "Metrics Collection Duration", "LLM Recommendation Duration", "Total Region Processing Time"
    ]
    
    curr.execute(f"""
        SELECT DISTINCT region FROM discovered_resources WHERE account_id = ? AND service_type = ? {region_clause}
        UNION
        SELECT DISTINCT region FROM billing_service_cache WHERE account_id = ? AND service_name = ? {region_clause}
    """, params + params)
    regions = [r['region'] for r in curr.fetchall()]
    
    if not regions:
        create_empty_state_message(ws2, headers2)
    else:
        apply_header_formatting(ws2, headers2)
        for reg in regions:
            curr.execute("SELECT COUNT(resource_id) FROM discovered_resources WHERE account_id = ? AND service_type = ? AND region = ?", (account_id, service_name, reg))
            r_disc = curr.fetchone()[0]
            
            curr.execute("SELECT recommendation FROM resource_summaries WHERE account_id = ? AND service_type = ? AND region = ?", (account_id, service_name, reg))
            r_recs = curr.fetchall()
            r_analysed = len(r_recs)
            r_gen = 0
            r_no_met = 0
            for r in r_recs:
                if r['recommendation'] in ("Unknown", "Analysis Failed"):
                    r_no_met += 1
                else:
                    r_gen += 1
                    
            reg_status = "Completed Successfully" if r_no_met == 0 else ("Partial Success" if r_gen > 0 else "Analysis Failed/Missing Metrics")
            
            ws2.append([
                reg, reg_status, pl_status, r_disc, r_analysed, r_gen, r_no_met,
                "N/A (Not Persisted)", "N/A", "N/A", "N/A"
            ])
        auto_size_columns(ws2)


    # --- Sheet 3: Resource Summary ---
    ws3 = wb.create_sheet(title="Resource Summary")
    headers3 = [
        "Resource ID", "Resource Name", "Resource Type", "Region", 
        "Current Configuration", "Discovery Status", "Metrics Fetch Status", 
        "Analysis Status", "Deterministic Metrics Summary"
    ]
    
    curr.execute(f"""
        SELECT d.resource_id, d.region, d.resource_type, d.metadata_json, 
               s.summary_json, s.recommendation
        FROM discovered_resources d
        LEFT JOIN resource_summaries s ON d.resource_id = s.resource_id AND d.account_id = s.account_id
        WHERE d.account_id = ? AND d.service_type = ? {region_clause.replace('region', 'd.region')}
    """, params)
    all_res = curr.fetchall()
    
    if not all_res:
        create_empty_state_message(ws3, headers3)
    else:
        apply_header_formatting(ws3, headers3)
        for res in all_res:
            res_id = res['resource_id']
            region = res['region']
            res_type = res['resource_type']
            meta = json.loads(res['metadata_json'])
            res_name = meta.get("name", meta.get("function_name", "N/A"))
            curr_config = res_type
            
            discovery_status = "Success"
            metrics_fetch_status = "Success" if res['summary_json'] and json.loads(res['summary_json']).get("metrics") else "Failed/No Metrics"
            analysis_status = res['recommendation'] if res['recommendation'] else "Pending"
            
            det_summary = ""
            if res['summary_json']:
                s_json = json.loads(res['summary_json'])
                metrics = s_json.get("metrics", {})
                if metrics:
                    summary_parts = []
                    for m_name, m_data in metrics.items():
                        summary_parts.append(f"{m_name}: Avg {m_data.get('average', 0):.2f}, Max {m_data.get('maximum', 0):.2f}")
                    det_summary = " | ".join(summary_parts)
                else:
                    det_summary = "No metrics data."
            else:
                det_summary = "Not analyzed yet."
                
            ws3.append([
                res_id, res_name, res_type, region, curr_config, 
                discovery_status, metrics_fetch_status, analysis_status, det_summary
            ])
        auto_size_columns(ws3)


    # --- Sheet 4: AI Recommendations ---
    ws4 = wb.create_sheet(title="AI Recommendations")
    
    headers4a = [
        "Resource ID", "Resource Name", "Region", "Current Configuration", 
        "Recommendation", "Suggested Configuration", "Confidence", "AI Reasoning", "AI Summary"
    ]
    apply_header_formatting(ws4, headers4a)
    
    has_section_a = False
    
    for res in all_res:
        rec = res['recommendation']
        if rec and rec not in ("Unknown", "Analysis Failed"):
            res_id = res['resource_id']
            meta = json.loads(res['metadata_json'])
            res_name = meta.get("name", meta.get("function_name", "N/A"))
            region = res['region']
            curr_config = res['resource_type']
            
            base_rec = "Keep Current"
            sugg_config = "N/A"
            if "Upsize" in rec:
                base_rec = "Upsize"
                if "(" in rec: sugg_config = rec.split("(")[1].replace(")", "")
            elif "Downsize" in rec:
                base_rec = "Downsize"
                if "(" in rec: sugg_config = rec.split("(")[1].replace(")", "")
            elif "Recommend Specific" in rec:
                base_rec = "Recommend Specific Instance"
                if "(" in rec: sugg_config = rec.split("(")[1].replace(")", "")
                
            if rec_filter and rec_filter != "All Recommendations":
                if base_rec != rec_filter:
                    continue
                    
            has_section_a = True
            
            reasoning = ""
            if res['summary_json']:
                curr.execute("SELECT explanation FROM resource_summaries WHERE resource_id = ? AND account_id = ?", (res_id, account_id))
                exp_row = curr.fetchone()
                if exp_row:
                    reasoning = exp_row['explanation']
                    
            ws4.append([
                res_id, res_name, region, curr_config, base_rec, sugg_config, 
                "High", reasoning, reasoning
            ])
            
    if not has_section_a:
        ws4.append(["No recommendations match the selected filters."])
        
    ws4.append([])
    ws4.append(["Total Upsize", total_upsize])
    ws4.append(["Total Downsize", total_downsize])
    ws4.append(["Total Keep Current", total_keep])
    ws4.append([])
    
    ws4.append(["--- Section B: Resources Without Metrics ---"])
    headers4b = ["Resource ID", "Resource Name", "Region", "Reason"]
    ws4.append(headers4b)
    
    b_row_idx = ws4.max_row
    for cell in ws4[b_row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C0504D")
        
    has_section_b = False
    for res in all_res:
        rec = res['recommendation']
        if rec in ("Unknown", "Analysis Failed"):
            has_section_b = True
            res_id = res['resource_id']
            meta = json.loads(res['metadata_json'])
            res_name = meta.get("name", meta.get("function_name", "N/A"))
            region = res['region']
            
            reason = "No CloudWatch Metrics"
            if res['summary_json']:
                s = json.loads(res['summary_json'])
                if "No metrics available" in s.get("explanation", "") or not s.get("metrics"):
                    reason = "No CloudWatch Metrics"
            
            ws4.append([res_id, res_name, region, reason])
            
    if not has_section_b:
        ws4.append(["All resources successfully fetched metrics."])
        
    auto_size_columns(ws4)

    # --- Sheet 5: Service Inventory ---
    ws5 = wb.create_sheet(title="Service Inventory")
    headers5 = ["Service Name", "Inventory Data"]
    apply_header_formatting(ws5, headers5)
    ws5.append([service_name, "Service-specific inventory export will be implemented in the next phase using the approved Inventory Plugin architecture."])
    auto_size_columns(ws5)
    
    conn.close()
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return stream
